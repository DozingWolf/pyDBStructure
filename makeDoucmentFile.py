# coding=utf-8
# @Author  : Edmond

from re import sub
from openpyxl import Workbook,styles
from loguru import logger

@logger.catch
def createExcel(dbtype:str,dataset:list,path:str,filename:str):
    # create workbook
    savePath = ''.join([path,filename,'.xlsx'])
    excel = Workbook()
    # create index sheet
    indexSheet = excel.create_sheet('IndexPage',0)
    # define title
    indexSheet.merge_cells(start_row = 1,start_column = 1,end_row = 1, end_column = 3)
    indexTitle = indexSheet.cell(row = 1 ,column = 1)
    indexTitle.value = '表结构目录'
    indexTitle.font = styles.Font(size=15,color=styles.colors.BLACK,bold=True)
    indexTitle.fill = styles.PatternFill(fgColor=styles.colors.COLOR_INDEX[17])
    indexTitle.alignment = styles.Alignment(horizontal='center', vertical='center')
    
    # define db type
    indexDBTypeTitle = indexSheet.cell(row = 2 , column = 1)
    indexDBTypeTitle.value = 'DB TYPE:'
    indexDBTypeTitle.alignment = styles.Alignment(horizontal='left', vertical='center')
    indexDBTypeTitle.font = styles.Font(color=styles.colors.BLACK,bold=True)
    indexDBType = indexSheet.cell(row = 2 , column = 2)
    indexDBType.value = dbtype
    indexDBType.alignment = styles.Alignment(horizontal='right', vertical='center')
    indexDBType.font = styles.Font(color=styles.colors.BLACK,bold=True)

    # define index table
    indexTableTitleID = indexSheet.cell(row = 3 , column = 1)
    indexTableTitleID.value = 'Index'
    indexTableTitleID.alignment = styles.Alignment(horizontal='center', vertical='center')
    indexTableTitleTName = indexSheet.cell(row = 3 , column = 2)
    indexTableTitleTName.value = 'Object Name'
    indexTableTitleTName.alignment = styles.Alignment(horizontal='center', vertical='center')
    for id , table in enumerate(dataset):
        rowID = id + 4
        indexRowColumn = indexSheet.cell(row = rowID,column = 1)
        indexRowColumn.value = table.get('id')+1
        indexRowTableColumn = indexSheet.cell(row = rowID,column = 2)
        indexRowTableColumn.value = table.get('TableName')
    # create data sheet
    for id , table in enumerate(dataset):
        # logger.debug(table)
        # logger.debug(table.get('id'))
        # logger.debug(table.get('TableName'))
        tableName = table.get('TableName')
        tableIndex = table.get('id')
        sheetName = '_'.join(['Index',str(tableIndex+1)])
        # logger.debug(sheetName)
        newSheet = excel.create_sheet(sheetName,tableIndex+1)
        newSheet.merge_cells(start_row = 1,start_column = 1,end_row = 1, end_column = 11)
        titleString = newSheet.cell(row = 1,column = 1)
        indexString = newSheet.cell(row = 2,column = 1)
        titleString.value = ':'.join(['Table',tableName])
        titleString.alignment = styles.Alignment(horizontal='center', vertical='center')
        indexString.value = '.'.join(['Index',str(tableIndex+1)])
        # logger.debug(table.get('TableStructureData'))
        tableStrucData = table.get('TableStructureData')
        newSheet.append([])

        # add table structure info
        newSheet.append(['Table Structure'])
        newSheet.append(['ID','Name','Desc','IsIdentity(not for oracle)','PK','Type','Byte Length','Length','Scale','Nullable','Default'])
        # logger.debug(tableStrucData)
        for rowData in tableStrucData:
            rDataAsList = [x for x in rowData]
            # logger.debug('rDataAsList[2] is:')
            # logger.debug(rDataAsList[2])
            
            # replace unknow utf-8 charect
            if rDataAsList[2] is not None:
                subComment = sub(r'[\x00-\x1F\x7F]','',rDataAsList[2])
            else:
                subComment = rDataAsList[2]
            newSheet.append([rDataAsList[0],rDataAsList[1],subComment,rDataAsList[3],rDataAsList[4],
                             rDataAsList[5],rDataAsList[6],rDataAsList[7],rDataAsList[8],rDataAsList[9],
                             rDataAsList[10]])
        newSheet.append([])
        
        # add table constraint info
        newSheet.append(['Table Constraint'])
        tableConstraint = table.get('TableConstraint')
        # logger.debug(tableConstraint)
        newSheet.append(['ID','Constraint name','Column name','Constraint Type'])
        if len(tableConstraint) != 0: 
            for idx,constData in enumerate(tableConstraint):
                cDataAsList = [x for x in constData]
                newSheet.append([idx+1,cDataAsList[0],cDataAsList[1],cDataAsList[2]])
        else:
            newSheet.append(['/','/','/','/'])
        newSheet.append([])

        # add column index info
        newSheet.append(['Column Index'])
        tableColumnIndex = table.get('TableColumnIndex')
        newSheet.append(['ID','Index Name','Index Type','Uniqueness','Compression'])
        if len(tableColumnIndex) != 0:
            for idx,indexData in enumerate(tableColumnIndex):
                iDataAsList = [x for x in indexData]
                newSheet.append([idx+1,iDataAsList[0],iDataAsList[1],iDataAsList[2],iDataAsList[3]])
        else:
            newSheet.append(['/','/','/','/','/'])

    excel.save(savePath)
